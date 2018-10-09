import urllib.request, urllib.error, urllib.parse
from openpyxl import load_workbook

csvFilePath = "https://metrolab.amp.pt/wp-content/uploads/2018/08/1_CSV_caraterizacao_de_residuos_municipios_LIPOR.csv"
xlsFilePath = "https://metrolab.amp.pt/wp-content/uploads/2018/10/2_XLS_producao_de_residuos_municipios_LIPOR.xlsx"

csvFilePath = urllib.request.Request(csvFilePath)
xlsFilePath = urllib.request.Request(xlsFilePath)

csvFile = urllib.request.urlopen(csvFilePath).read()
xlsFile = urllib.request.urlopen(xlsFilePath).read()

with open("file.csv", "wb") as csvFileWrite:
    csvFileWrite.write(csvFile)

with open("file.csv", "r") as csv:
    csvFile = csv.readlines()

header = csvFile[0].strip().split(";")
dataList = []

for k in range(1, len(csvFile)):
    dataList.append(csvFile[k].strip().split(";"))

for k in header:
    print(k, end=" | ")
print("\r")
found = False

for line in range(len(dataList)):
    for column in range(len(dataList[line])):
        if (column == 0):
            print("\r")

        print(dataList[line][column], end=" | ")

with open("file.xlsx", "wb") as xlsFileWrite:
    xlsFileWrite.write(xlsFile)

wb = load_workbook("file.xlsx", read_only=True)
sheetName = wb.sheetnames[0]
ws = wb[sheetName]
for row in ws.rows:
    for cell in row:
        print(cell.value)
